"""
export_excel.py
Builds a downloadable .xlsx workbook of the routine grid(s) using openpyxl.

Layout matches the on-screen grid: days on rows, time slots on columns.
Supports per-semester time slots, merged (multi-period) assignments, and
break slots that may have assignments.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

import database as db

INK = "1F2A44"
GOLD = "B8863B"
CREAM = "FAF7F0"
BREAK_FILL = "EAE4D6"

thin = Side(style="thin", color="C9C2AE")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _sheet_name(program, semester):
    short = {"B.Tech Computer Science": "BTech-CS", "BSc Data Science": "BSc-DS", "BCA": "BCA"}.get(
        program, program
    )
    name = f"{short} Sem{semester}"
    return name[:31]  # Excel sheet name limit


def _compute_merge_map(slots, assignments):
    """Build horizontal merge map: ``'day|slot_id' -> span``

    - ``span > 1``: first cell of a horizontally-merged group.
    - ``span == -1``: cell swallowed by the merge to its left.
    """
    merge_map = {}
    for day in db.DAYS:
        i = 0
        while i < len(slots):
            key = f"{day}|{slots[i]['id']}"
            entry = assignments.get(key)
            if entry:
                span = 1
                while i + span < len(slots):
                    next_key = f"{day}|{slots[i + span]['id']}"
                    next_entry = assignments.get(next_key)
                    if next_entry and next_entry.get("subject_id") == entry.get("subject_id"):
                        span += 1
                    else:
                        break
                if span > 1:
                    merge_map[key] = span
                    for j in range(1, span):
                        merge_map[f"{day}|{slots[i + j]['id']}"] = -1
                i += span
            else:
                i += 1
    return merge_map


def build_workbook(program_semesters):
    """program_semesters: list of (program, semester) tuples, one sheet each."""
    wb = Workbook()
    wb.remove(wb.active)

    for program, semester in program_semesters:
        slots = db.list_time_slots(program, semester)
        if not slots:
            slots = []

        ws = wb.create_sheet(_sheet_name(program, semester))
        assignments = db.get_assignments(program, semester)
        merge_map = _compute_merge_map(slots, assignments)

        # Title row — spans day-label column + all slot columns.
        total_cols = 1 + len(slots)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=f"{program} — Semester {semester}")
        title_cell.font = Font(size=14, bold=True, color=INK)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # Header row: "Day" + one column per time slot.
        header_row = 2
        day_hdr = ws.cell(row=header_row, column=1, value="Day")
        day_hdr.font = Font(bold=True, color="FFFFFF")
        day_hdr.fill = PatternFill("solid", fgColor=INK)
        day_hdr.alignment = Alignment(horizontal="center")
        day_hdr.border = BORDER

        for ci, slot in enumerate(slots, start=2):
            slot_label = f"{slot['label']}\n{slot['start_time']}-{slot['end_time']}"
            c = ws.cell(row=header_row, column=ci, value=slot_label)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor=INK)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
        ws.row_dimensions[header_row].height = 36

        # Body rows: one row per day.
        r = header_row + 1
        for day in db.DAYS:
            # Day label cell (row header).
            dc = ws.cell(row=r, column=1, value=day)
            dc.font = Font(bold=True, color=INK, size=10)
            dc.alignment = Alignment(horizontal="center", vertical="center")
            dc.border = BORDER
            dc.fill = PatternFill("solid", fgColor=CREAM)

            for ci, slot in enumerate(slots, start=2):
                key = f"{day}|{slot['id']}"
                m = merge_map.get(key)
                entry = assignments.get(key)
                cell = ws.cell(row=r, column=ci)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if m == -1:
                    # Swallowed by a horizontal merge to its left.
                    continue

                if m and m > 1:
                    ws.merge_cells(
                        start_row=r, start_column=ci,
                        end_row=r, end_column=ci + m - 1,
                    )

                if entry:
                    cell.value = (
                        f"{entry['subject_code']}\n{entry['subject_name']}\n"
                        f"{entry['teacher_name'] or ''}"
                    )
                    cell.font = Font(size=9, color=INK)
                    cell.fill = PatternFill("solid", fgColor="EFEBDD")
                elif slot["is_break"]:
                    cell.value = slot["label"]
                    cell.font = Font(italic=True, color="6B6350")
                    cell.fill = PatternFill("solid", fgColor=BREAK_FILL)
                else:
                    cell.value = ""

            r += 1

        # Column widths.
        ws.column_dimensions["A"].width = 14
        for ci in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 16
        for row in ws.iter_rows(min_row=header_row + 1, max_row=r - 1):
            ws.row_dimensions[row[0].row].height = 50
        ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
